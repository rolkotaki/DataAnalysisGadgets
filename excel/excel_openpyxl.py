import pandas as pd
from openpyxl import Workbook
from openpyxl import load_workbook
from openpyxl.styles import Font, colors, Color, Alignment, PatternFill, GradientFill, Border, Side, NamedStyle
from openpyxl.chart import PieChart, PieChart3D, Reference, Series, BarChart, LineChart, AreaChart
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.image import Image
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.comments import Comment
from openpyxl.worksheet.datavalidation import DataValidation


# https://openpyxl.readthedocs.io/en/stable/
# https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.read_excel.html
# https://pandas.pydata.org/pandas-docs/stable/reference/api/pandas.DataFrame.to_excel.html

# Beside openpyxl there is another Python library for Excel: https://xlsxwriter.readthedocs.io/


# Creating workbooks and worksheets
wb = Workbook()  # to create an empty Workbook object
ws1 = wb.create_sheet('London')
ws2 = wb.create_sheet('Madrid', 0)  # creating a new sheet; parameters: sheet name + index

# Worksheet operations
wb.copy_worksheet(ws1)

wb.create_sheet('Delete')
ws = wb['Delete']
wb.remove(ws)  # deleting a worksheet

# Setting active worksheet
wb.active = wb['Madrid']
ws = wb.active

# Dummy data
for i in range(1, 20):     # looping through the first 19 rows
    ws.append(range(300))  # filling the first 300 column with values from 0 to 299

# Rows and columns
ws.delete_rows(15, 2)  # deleting 2 rows starting from 15
ws.delete_cols(17, 3)  # deleting columns Q, R and S

ws.insert_rows(15, 2)  # inserting two rows at the 15th row
ws.insert_cols(17, 3)  # inserting three columns at the 17th column

# Moving a range
ws.move_range("F19:H19", rows=2, cols=2)  # translate=True

# Selecting cells
cell_range = ws['A1':'C1']
col_range = ws['C']   # selecting the whole column
row_range = ws[5:10]  # selecting the whole rows

# Merging cells
ws.merge_cells('A1:B5')
ws.unmerge_cells('A1:B5')
ws.merge_cells(start_row=2, start_column=2, end_row=5, end_column=5)

# Working with cells
cell = ws['B2']
cell.value = 'Merged cell'
cell.font = Font(color=colors.RED, size=20, italic=True, bold=True)
cell.alignment = Alignment(horizontal='left', vertical='top')
cell.fill = GradientFill(stop=('000000', 'FFFFFF'))  # black and white, hexa code for RGB

# Creating a custom style that we can use later
highlight = NamedStyle(name='highlight')
highlight.font = Font(bold=True)
border = Side(style='thick', color='000000')  # creating border style options
highlight.border = Border(left=border, top=border, right=border, bottom=border)
highlight.fill = PatternFill('solid', fgColor='FFFF00')

# iterating through columns using the openpyxl's iterator function:
row_num = 0
for col in ws.iter_cols(min_col=8, max_col=15, min_row=1, max_row=20):
    col[row_num].style = highlight  # setting the cell style to our customized style
    # row_num += 1                    # we increase the row number in each loop, it will look like a diagonal

# the same way we can iterate through rows:
# values_only: if True, we only cell values are looped through. if False, cells as objects
for row in ws.iter_rows(min_row=1, max_row=2, max_col=3, values_only=True):
    for cell in row:
        print(cell)  # if values_only is False, then: print(cell.value), because then objects are returned, not values

# Cell comments
comment = Comment(text='This is a comment', author='Darth Vader')
ws['A10'].comment = comment

# Cell formula
ws['A11'] = '=SUM(1, 1)'

# Cell format
ws['A12'].number_format = '0000000000000'

# Charts
data_list = [['Capital', 'SomeInfo'],
             ['London', 1000],
             ['Budapest', 999],
             ['Madrid', 850],
             ['Paris', 2900],
             ['New York', 5000]]

ws = wb['London']

for row in data_list:
    ws.append(row)

labels = Reference(ws, min_row=2, min_col=1, max_row=6)
data = Reference(ws, min_row=2, min_col=2, max_row=6)

chart = PieChart()
chart.add_data(data)
chart.set_categories(labels)
chart.title = 'Capitals Dummy Information - PieChart'
ws.add_chart(chart, 'D1')

chart = LineChart()
chart.add_data(data)
chart.set_categories(labels)
chart.title = 'Capitals Dummy Information - LineChart'
ws.add_chart(chart, 'M1')

chart = BarChart()
chart.add_data(data)
chart.set_categories(labels)
chart.title = 'Capitals Dummy Information - BarChart'
ws.add_chart(chart, 'D17')

chart = AreaChart()
chart.add_data(data)
chart.set_categories(labels)
chart.title = 'Capitals Dummy Information - AreaChart'
ws.add_chart(chart, 'M17')

# Tables
# Let's create a table for our previous data
table = Table(displayName='Capitals', ref='A1:B6')
style = TableStyleInfo(name='TableStyleMedium9', showFirstColumn=False, showLastColumn=False,
                       showRowStripes=True, showColumnStripes=True)
table.tableStyleInfo = style
ws.add_table(table)

# Filtering
ws = wb['Sheet']
for row in data_list:
    ws.append(row)
ws.auto_filter.ref = 'A1:B1'

# Freezing
ws.freeze_panes = ws['B2']

# Data Validation
# https://openpyxl.readthedocs.io/en/stable/validation.html
dv = DataValidation(type="list", formula1='"Dog,Cat,Bat"', allow_blank=True)
ws.add_data_validation(dv)
cell = ws['A10']
dv.add(cell)

# Images
wb.create_sheet('Image')
img = Image('../data/Madrid.jpg')
img.height *= 0.2
img.width *= 0.2
wb['Image'].add_image(img, 'C5')

# Saving the workbook
wb.save('../data/test.xlsx')

# Loading an existing workbook
wb_shifts = load_workbook('../data/regions.xlsx')

# Pivot tables
# https://openpyxl.readthedocs.io/en/stable/pivot.html
# https://openpyxl.readthedocs.io/en/stable/api/openpyxl.pivot.cache.html#openpyxl.pivot.cache.CacheDefinition


# ********** Excel with pandas

# Loading an Excel sheet into a dataframe
df = pd.read_excel('../data/shifts.xlsx', sheet_name='Sheet',
                   header=0,  # headers in the first row
                   skiprows=None,  # don't skip any rows
                   names=None,  # if there are no headers in the file, we can specify the column names
                   index_col=None,  # we could specify the column(s) to be the index; one or more column indexes
                   usecols=None,  # None: all columns; str: comma separated Excel column letters;
                                  # int: list of column numbers; str list: list of column names
                   dtype=None,  # please check file_management/df_save_load_csv.py file, same as for CSV
                   parse_dates=False  # whether to parse dates or not (if not, they will be loaded as plain objects)
                   # na_values, keep_default_na, na_filter
                   )
# Most of the parameters behave the same way as shown in file_management/df_save_load_csv.py file for CSV files.

# we can use openpyxl's iterator function to iterate through the dataframe (worksheet data)
rows = dataframe_to_rows(df, header=False, index=False)
for row_idx, row in enumerate(rows, 1):     # start indexing from 1
    for col_idx, col in enumerate(row, 1):  # start indexing from 1
        pass  # we can do whatever

# Saving a dataframe into an Excel  sheet
df.to_excel('../data/exported_excel.xlsx',
            sheet_name='Exported',
            na_rep='',  # for empty values, default is an empty string
            # float_format
            columns=['Region', 'Sales Rep', 'Product', 'Cost per', 'Units Sold'],  # columns to be exported
            header=True,
            index=True,  # whether to write index in the file or not
            index_label=['Index'],
            startrow=0,
            startcol=0
            # engine: 'openpyxl' or 'xlsxwriter'
            # merge_cells: default True
            )

# If we have to perform operations on the data, then it may be easier to load the Excel Sheet data into a pandas
# dataframe, do all the operations, and then export it back.
